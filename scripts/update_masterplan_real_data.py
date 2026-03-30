"""Update frai_tv_masterplan.xlsx with real channel data (396 videos)"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

wb = load_workbook('frai_tv_masterplan.xlsx')
nf = Font(name='Arial', size=10)
gf = Font(name='Arial', bold=True, size=11, color='D4AF37')
ff = Font(name='Arial', size=10, color='000000')
rf = Font(name='Arial', size=10, color='CC0000', bold=True)
gold_fill = PatternFill('solid', fgColor='D4AF37')
green_f = PatternFill('solid', fgColor='E8F5E9')
red_f = PatternFill('solid', fgColor='FFEBEE')
yellow_f = PatternFill('solid', fgColor='FFF8E1')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

# === UPDATE SHEET 1: Real numbers ===
ws1 = wb['Kanal_Uebersicht']
ws1.cell(row=7, column=2, value=396)  # Videos YouTube
ws1.cell(row=7, column=3, value='Verifiziert via YouTube Studio Screenshot')
ws1.cell(row=8, column=2, value=92)   # remaikeData.js
ws1.cell(row=8, column=3, value='NUR 92 von 396! 304 fehlen!')
ws1.cell(row=8, column=3).font = rf

# Add real channel stats
r = 24
ws1.cell(row=r, column=1, value='ECHTE KANAL-STATISTIKEN').font = gf
ws1.cell(row=r, column=1).fill = gold_fill
for c in range(1, 4):
    ws1.cell(row=r, column=c).border = tb

real_stats = [
    ('Subscriber', '22.400', 'Social Blade verifiziert'),
    ('Gesamtvideos', 396, 'YouTube Studio 21.03.2026'),
    ('Upload-Frequenz', '~14/Woche', 'Sehr aktiv'),
    ('Avg. Videolaenge', '21 Min', ''),
    ('Est. monatl. Einnahmen', '$8-$25', 'Social Blade'),
    ('Views letzte 7 Tage', '+1.494', '+46.18%'),
    ('', '', ''),
    ('TOP PERFORMER', 'Views', 'Faktor'),
    ('White Zombie (1932)', '32.812', '21.4x Durchschnitt'),
    ('Felix the Cat (1921)', '25.077', '>100x Outlier'),
    ('Superman (1941)', '20.295', '11x'),
    ('Dr. Caligari (1920)', '16.621', '12.4x'),
    ('Dinner for One (1963)', '14.226', '38.9x'),
    ('Ken Block Gymkhana', '8.686', '>100x'),
    ('Charade (1953)', '7.722', '5.8x'),
    ('Grim Game Houdini', '6.761', '2.6x'),
    ('', '', ''),
    ('WOCHENSCHAU Performance', 'Views', 'Problem'),
    ('Wochenschau 459: Pre-War', '1.948', 'Einziger Outlier'),
    ('Wochenschau 511: Paris', '155', 'Niedrig'),
    ('Wochenschau 515: Battle of Britain', '34', 'Sehr niedrig'),
    ('Durchschnitt Wochenschau', '~50-80', 'Titel-Format killt CTR'),
    ('', '', ''),
    ('CONTENT-VERTEILUNG (342 verifiziert)', 'Anzahl', 'Serie'),
    ('Betty Boop', 57, 'Groesste Serie'),
    ('Deutsche Wochenschau', 45, 'WWII Newsreels'),
    ('Alfred J. Kwak', 34, 'Anime DE'),
    ('Soundies (1940s Musik)', 33, 'Vintage Music'),
    ('Felix the Cat', 13, 'Stummfilm-Cartoons'),
    ('Superman', 10, 'Fleischer Studios'),
    ('Casper', 9, 'Friendly Ghost'),
    ('Porky Pig', 8, 'Looney Tunes'),
    ('Krtek (Maulwurf)', 7, 'Tschechisch'),
    ('Teaserama', 7, 'Burlesque'),
    ('Ken Block', 4, 'Automotive'),
    ('Popeye', 4, 'Famous Studios'),
    ('Astro Boy', 3, 'Anime'),
    ('Looney Tunes', 3, 'Warner Bros'),
    ('BraveStarr', 2, 'TV-Serie'),
    ('Asterix', 2, 'Animation'),
    ('Sonstige (Filme, Dokus etc.)', '~102', 'Diverse'),
    ('', '', ''),
    ('FEHLENDE VIDEOS', '', ''),
    ('In remaikeData.js', 92, ''),
    ('Auf YouTube', 396, ''),
    ('FEHLEND IM CODE', '=B64-B63', 'KRITISCH! Formel'),
    ('In Sitemap/API gefunden', 342, ''),
    ('Noch nicht gefunden', '=B64-B66', 'Wahrsch. Shorts/Neueste'),
]

for i, (label, value, note) in enumerate(real_stats, r + 1):
    ws1.cell(row=i, column=1, value=label).font = nf
    ws1.cell(row=i, column=1).border = tb
    c = ws1.cell(row=i, column=2, value=value)
    c.font = ff if str(value).startswith('=') else nf
    c.border = tb
    ws1.cell(row=i, column=3, value=note).font = nf
    ws1.cell(row=i, column=3).border = tb
    if 'KRITISCH' in str(note):
        ws1.cell(row=i, column=3).font = rf
    if 'TOP PERFORMER' in str(label) or 'WOCHENSCHAU' in str(label) or 'CONTENT-VERTEILUNG' in str(label) or 'FEHLENDE' in str(label):
        ws1.cell(row=i, column=1).font = gf

# === ADD NEW SHEET: Praesentations_Analyse ===
ws_pa = wb.create_sheet('Praesentations_Analyse', 2)
ws_pa.sheet_properties.tabColor = 'FF6F00'

hdr_fill = PatternFill('solid', fgColor='2B2B2C')
hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')

pa_headers = ['Faktor', 'Top Performer (White Zombie etc.)', 'Wochenschau (34 Views)', 'Empfehlung', 'Impact']
for c, h in enumerate(pa_headers, 1):
    cl = ws_pa.cell(row=1, column=c, value=h)
    cl.font = hf
    cl.fill = hdr_fill
    cl.alignment = Alignment(horizontal='center', wrap_text=True)
    cl.border = tb

ws_pa.column_dimensions['A'].width = 22
ws_pa.column_dimensions['B'].width = 40
ws_pa.column_dimensions['C'].width = 40
ws_pa.column_dimensions['D'].width = 45
ws_pa.column_dimensions['E'].width = 12

pa_data = [
    ('TITEL-FORMAT',
     'White Zombie (1932) | Horror | 8K',
     'Wochenschau 515: Battle of Britain (Jul 1940) | 8K',
     'Event ZUERST: Battle of Britain (1940) | Deutsche Wochenschau | 8K',
     'KRITISCH'),
    ('THUMBNAIL',
     'Erkennbares Gesicht (Bela Lugosi), ikonisch',
     'Graues Militaer-Footage, alle sehen gleich aus',
     'Kolorierte Key-Frames, Event-Name gross, differenzieren',
     'KRITISCH'),
    ('SUCHBEGRIFFE',
     '"white zombie full movie" = kaum 8K Konkurrenz',
     '"battle of britain" = Netflix/Prime/BBC dominieren',
     'Nischen-Keywords: "WWII German newsreel 8K original"',
     'Hoch'),
    ('CONTENT-LAENGE',
     '65-90 Min (Full Movie) = hohe Watchtime',
     '15-25 Min (Newsreel) = kurze Sessions',
     'Kompilationen: "Complete German Newsreel 1940 | 2h" ',
     'Hoch'),
    ('WIEDERERKENNUNGSWERT',
     'Superman, Felix, Chaplin = globale Brands',
     '"Wochenschau Nr. 515" = Datenbank-Dump-Look',
     'Serien-Branding: "WWII Archives" statt Nummern',
     'Hoch'),
    ('BESCHREIBUNG',
     'Kurz, praegnant, Genre-Keywords',
     'Multi-Language OK, aber zu generisch',
     'Hook in Zeile 1, historischer Kontext, Chapters',
     'Mittel'),
    ('HASHTAGS',
     '#ClassicFilm #Horror #8K #Remastered',
     '#Wochenschau #WWII #History #8K',
     'EN Tags: #WWII #Documentary #BattleOfBritain #8K',
     'Mittel'),
    ('ZIELGRUPPE',
     'Filmfans, Nostalgie, Horror-Community',
     'Historiker, WWII-Enthusiasten (kleinere Nische)',
     'EN-Titel fuer globale Reichweite, DE als Untertitel',
     'Hoch'),
    ('SERIE/PLAYLIST',
     'Einzelne Filme (standalone)',
     'Teil einer 45+ Serie (aber keine Playlist!)',
     'Thematische Playlists: "Eastern Front", "Western Front"',
     'Hoch'),
    ('UPLOAD-KONTEXT',
     'Aeltere Videos, mehr Zeit fuer Discovery',
     'Neueste Uploads (Feb 2026), noch frisch',
     'Geduld + Optimierung = organisches Wachstum',
     'Niedrig'),
]

for i, row in enumerate(pa_data, 2):
    for j, val in enumerate(row, 1):
        c = ws_pa.cell(row=i, column=j, value=val)
        c.font = nf
        c.border = tb
        c.alignment = Alignment(wrap_text=True)
        if val == 'KRITISCH':
            c.font = rf
            c.fill = red_f
        elif val == 'Hoch':
            c.font = Font(name='Arial', size=10, color='CC6600', bold=True)
            c.fill = yellow_f

# Summary row
sr = len(pa_data) + 3
ws_pa.cell(row=sr, column=1, value='FAZIT').font = gf
ws_pa.cell(row=sr, column=1).fill = gold_fill
ws_pa.merge_cells(f'A{sr}:E{sr}')
ws_pa.cell(row=sr+1, column=1, value='Die 2 kritischsten Faktoren sind TITEL-FORMAT und THUMBNAIL.').font = nf
ws_pa.cell(row=sr+2, column=1, value='Titel-Aenderung allein kann CTR um 50-200% steigern (YouTube Research).').font = nf
ws_pa.cell(row=sr+3, column=1, value='Empfehlung: Alle 45 Wochenschau-Titel umformatieren + individuelle Thumbnails.').font = rf

wb.save('frai_tv_masterplan.xlsx')
print('XLSX aktualisiert mit echten Kanal-Daten + Praesentations-Analyse!')
print(f'Sheets: {wb.sheetnames}')
