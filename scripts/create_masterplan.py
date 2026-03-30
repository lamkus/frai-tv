from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
hdr_fill = PatternFill('solid', fgColor='2B2B2C')
gold_fill = PatternFill('solid', fgColor='D4AF37')
dark_fill = PatternFill('solid', fgColor='1A1A1B')
green_f = PatternFill('solid', fgColor='E8F5E9')
red_f = PatternFill('solid', fgColor='FFEBEE')
yellow_f = PatternFill('solid', fgColor='FFF8E1')
hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
gf = Font(name='Arial', bold=True, size=11, color='D4AF37')
tf = Font(name='Arial', bold=True, size=14, color='D4AF37')
nf = Font(name='Arial', size=10)
ff = Font(name='Arial', size=10, color='000000')
rf = Font(name='Arial', size=10, color='CC0000', bold=True)
of = Font(name='Arial', size=10, color='FF6600')
tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def sh(ws, r, n):
    for c in range(1, n + 1):
        cl = ws.cell(row=r, column=c)
        cl.font = hf
        cl.fill = hdr_fill
        cl.alignment = Alignment(horizontal='center', wrap_text=True)
        cl.border = tb

# === SHEET 1: Kanal_Uebersicht ===
ws1 = wb.active
ws1.title = 'Kanal_Uebersicht'
ws1.sheet_properties.tabColor = 'D4AF37'
ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 30
ws1.merge_cells('A1:C1')
ws1['A1'] = 'remAIke.TV / frai.tv - Masterplan Dashboard'
ws1['A1'].font = tf
ws1['A1'].fill = dark_fill

ws1['A2'] = 'Metrik'
ws1['B2'] = 'Wert'
ws1['C2'] = 'Anmerkung'
sh(ws1, 2, 3)

overview = [
    ('Kanal-Name', '@remAIke_IT', 'YouTube Handle'),
    ('Channel-ID', 'UCVFv6Egpl0LDvigpFbQXNeQ', ''),
    ('Website', 'https://frai.tv', 'Companion Website'),
    ('Beschreibung', 'AI-restaurierte Public Domain Filme 4K/8K', ''),
    ('Videos YouTube (ca.)', 85, 'API-Check laeuft'),
    ('Videos remaikeData.js', 92, 'Code-Katalog'),
    ('davon Shorts', 12, 'isShort: true'),
    ('davon Long-Form', '=B8-B9', 'Formel: Total-Shorts'),
    ('Video-Seiten frai.tv', 114, 'Sitemap'),
    ('Schema.org Items', 406, 'ItemList Homepage'),
    ('Delta frai.tv vs YT', '=B11-B7', 'Formel'),
    ('Delta Schema vs Code', '=B12-B8', 'Formel'),
    ('Sprachen', 'DE / EN / FR', 'i18next'),
    ('Frontend', 'React 18 + Vite 5 + Tailwind', ''),
    ('Backend', 'Express + Prisma + PostgreSQL', ''),
    ('Hosting', 'Strato VPS Deutschland', ''),
    ('CDN', 'Keiner (Cloudflare geplant)', ''),
    ('Analytics', 'Geplant: stats.frai.tv Matomo', ''),
    ('CI/CD', 'GitHub Actions', 'ci.yml + deploy.yml'),
    ('PWA', 'Ja', 'manifest.json + Service Worker'),
]

for i, (label, value, note) in enumerate(overview, 3):
    ws1.cell(row=i, column=1, value=label).font = nf
    ws1.cell(row=i, column=1).border = tb
    c = ws1.cell(row=i, column=2, value=value)
    c.font = nf
    c.border = tb
    ws1.cell(row=i, column=3, value=note).font = nf
    ws1.cell(row=i, column=3).border = tb

# === SHEET 2: Video_Katalog ===
ws2 = wb.create_sheet('Video_Katalog')
ws2.sheet_properties.tabColor = 'E50914'

h2 = ['Nr', 'ID', 'ytId', 'Titel', 'Jahr', 'Kategorie', 'Subcategory', 'Dauer_Sek', 'Tags', 'isShort', 'Status_YT', 'Chapters', 'Playlist', 'Titel_OK', 'Beschr_OK']
for c, h in enumerate(h2, 1):
    ws2.cell(row=1, column=c, value=h)
sh(ws2, 1, len(h2))
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 14
ws2.column_dimensions['D'].width = 48
ws2.column_dimensions['E'].width = 6
ws2.column_dimensions['F'].width = 14
ws2.column_dimensions['G'].width = 14
ws2.column_dimensions['H'].width = 10
ws2.column_dimensions['I'].width = 35
ws2.column_dimensions['J'].width = 8

vids = [
    ('rudolph-1948', 'YzvGVo8mAQM', 'Rudolph the Red-Nosed Reindeer (1948)|8K', 1948, 'christmas', 'cartoons', 480, 'weihnachten,rudolph', False),
    ('suzy-snowflake', 'Z4nwcfOqOOw', 'Suzy Snowflake (1953)|8K', 1953, 'christmas', 'cartoons', 300, 'weihnachten,stop-motion', False),
    ('coca-cola-4k', 'U-WD47NSgAE', 'Coca-Cola Christmas Trucks (1995)|4K', 1995, 'commercials', 'coca-cola', 60, 'coca-cola,weihnachten', False),
    ('coca-cola-8k', 'WSjkAZkPbKs', 'EPIC Coca-Cola Christmas Trucks|8K', 1995, 'commercials', 'coca-cola', 60, 'coca-cola,weihnachten', False),
    ('batman-santa', 'yIQCHpjp4NE', 'Batman & Robin Christmas (1966)|8K', 1966, 'christmas', 'movies', 1500, 'batman,robin,santa', False),
    ('christmas-carol', 'dGD2CeoZX68', 'A Christmas Carol (1984)|8K', 1984, 'christmas', 'movies', 6000, 'dickens,scrooge', False),
    ('casper-xmas', 'uWacVV7EkxQ', 'Casper Christmas Special|8K', 1950, 'christmas', 'cartoons', 420, 'casper,weihnachten', False),
    ('great-expect', 'JhJmasQ8N-8', 'Great Expectations (1946)|8K', 1946, 'classic-films', 'drama', 7200, 'david lean,dickens', False),
    ('scarlet-st', '_aUNgDJoWoU', 'Scarlet Street (1945)|8K', 1945, 'classic-films', 'film-noir', 6300, 'fritz lang,film noir', False),
    ('metropolis', '8lLtNb11NKU', 'Metropolis (1927)|4K 8K', 1927, 'classic-films', 'silent', 9000, 'fritz lang,sci-fi', False),
    ('phantom', 'b4cqLlJ7t4M', 'Phantom of Opera (1925)|8K', 1925, 'classic-films', 'horror', 5400, 'lon chaney,horror', False),
    ('nosferatu', 'Nzi6aRKDoEs', 'Nosferatu (1922)|8K', 1922, 'classic-films', 'horror', 5400, 'murnau,vampire', False),
    ('tarzan', 'cCNm8nNHing', 'Tarzans Revenge (1938)|8K', 1938, 'classic-films', 'adventure', 4200, 'tarzan', False),
    ('20000-leagues', 'LEM6FkBTDNs', '20000 Leagues (1916)|8K', 1916, 'classic-films', 'silent', 6300, 'jules verne', False),
    ('haxan', 'exukLdxugy8', 'Haexan (1922)|8K', 1922, 'classic-films', 'horror', 6600, 'horror,witchcraft', False),
    ('frankenstein', 'hPQN992PMUY', 'Frankenstein (1910)|8K', 1910, 'classic-films', 'horror', 780, 'frankenstein,edison', False),
    ('voyage-preh', 'pqrCPhUCpxE', 'Voyage Prehistoric Women|8K', 1968, 'classic-films', 'adventure', 4800, 'sci-fi,cult', False),
    ('bill-divorce', 'YbC2JynVCRA', 'Bill of Divorcement (1932)|8K', 1932, 'classic-films', 'drama', 4200, 'hepburn', False),
    ('dinner-one', 'z8FqTSpp6Kg', 'Dinner for One (1963)|4K 8K', 1963, 'comedy', 'sketch', 1080, 'silvester,comedy', False),
    ('chaplin-fest', 'FG-vqRH5Cg4', 'Charlie Chaplin Film Fest|8K', 1917, 'comedy', 'chaplin', 5400, 'chaplin,marathon', False),
    ('chaplin-cure', 'EM076HMwVwI', 'The Cure (1917) Chaplin|8K', 1917, 'comedy', 'chaplin', 1500, 'chaplin', False),
    ('keaton-conv', '_3Z1GTYFUAM', 'Keaton Convict 13 (1920)|8K', 1920, 'comedy', 'keaton', 1200, 'keaton,slapstick', False),
    ('keaton-wife', 'mybF4jPjl64', 'Keaton My Wifes Relations|8K', 1922, 'comedy', 'keaton', 1200, 'keaton', False),
    ('lucy-wayne', 'AyjJbgijx68', 'Lucy Meets John Wayne (1966)|8K', 1966, 'comedy', 'lucy', 1500, 'lucy,john wayne', False),
    ('skeleton-d', 'ezYIk8bReaE', 'Skeleton Dance (1929)|8K', 1929, 'cartoons', 'disney', 360, 'disney,halloween', False),
    ('casper-boos', '3ePy5Hhq5sE', 'Casper Boos Saddles (1950)|8K', 1950, 'cartoons', 'casper', 420, 'casper,ghost', False),
    ('casper-ghost', 'eT-sQaNwF_w', 'Casper Once Ghostly (1949)|8K', 1949, 'cartoons', 'casper', 420, 'casper', False),
    ('peter-wolf', 'V8qeIRongA0', 'Peter and Wolf (1960)|8K', 1960, 'cartoons', 'fleischer', 600, 'prokofiev', False),
    ('superman-jap', 'bShrsrrzOYQ', 'Superman Japoteurs (1942)|8K', 1942, 'cartoons', 'superman', 540, 'superman,wwii', False),
    ('superman-elec', 'Aq1OfWdwV-Q', 'Superman Electric Earthquake|8K', 1942, 'cartoons', 'superman', 540, 'superman', False),
    ('superman-dest', 'e0Tagj2Z5SU', 'Superman Destruction Inc|8K', 1943, 'cartoons', 'superman', 540, 'superman', False),
    ('popeye-anc', 'unbwEeI4bEE', 'Popeye Ancient Fistory|8K', 1953, 'cartoons', 'popeye', 420, 'popeye', False),
    ('popeye-patr', 'eeId9wqhtuQ', 'Patriotic Popeye (1957)|8K', 1957, 'cartoons', 'popeye', 420, 'popeye', False),
    ('betty-boop', 'dkxCgOivonc', 'Betty Boop More Pep (1936)|8K', 1936, 'cartoons', 'betty-boop', 420, 'betty boop,jazz', False),
    ('little-nemo', 'T8AnrW3H5i8', 'Little Nemo (1911)|8K World First', 1911, 'cartoons', 'fleischer', 720, 'winsor mccay', False),
    ('kirby', 'Qm3K0-XL46Q', 'Kirby Abridged Collection|8K', 2005, 'cartoons', 'fleischer', 1800, 'kirby,anime', False),
    ('cigar-box', 'PySMbbpIAUM', 'Cigar Box (1905)|8K', 1905, 'classic-films', 'silent', 180, 'pathe', False),
    ('golden-beetle', 'IpWrF7DYyWU', 'Golden Beetle (1907)|8K', 1907, 'classic-films', 'silent', 240, 'pathe', False),
    ('frog-1908', 'dfUyhjEnAqw', 'The Frog (1908)|8K', 1908, 'classic-films', 'silent', 180, 'pathe', False),
    ('bee-rose', 'gMpONbJ3-9U', 'Bee and Rose (1908)|8K', 1908, 'classic-films', 'silent', 180, 'pathe', False),
    ('berlin-sym', 'DnwDqCsSqRw', 'Berlin Symphony (1927)|8K', 1927, 'documentaries', 'historical', 3900, 'berlin,weimar', False),
    ('pearl-harbor', 'EasEhYlorqQ', 'Pearl Harbor (1942)|8K', 1942, 'propaganda', 'wwii', 1800, 'wwii', False),
    ('bio-warfare', 'Zu_iBCd5NJc', 'Biological Warfare (1952)|8K', 1952, 'propaganda', 'civil-defense', 900, 'cold war', False),
    ('duck-cover', 'Ge-mC5q6lXw', 'Duck and Cover (1951)|8K', 1951, 'propaganda', 'civil-defense', 540, 'cold war', False),
    ('nazi-camps', 'DO8dSN4aAB4', 'Nazi Concentration Camps|8K', 1945, 'documentaries', 'historical', 3600, 'wwii,holocaust', False),
    ('my-japan', 'rV78L39ybOU', 'My Japan (1945) WWII|8K', 1945, 'propaganda', 'wwii', 600, 'wwii,propaganda', False),
    ('atomic-bomb', '9AgSJyMnxi8', 'Atomic Bomb Newsreel|Wochenschau', 1946, 'documentaries', 'newsreels', 420, 'wochenschau,atomic', False),
    ('hindenburg', 'eF81rBeXbzk', 'Hindenburg Explodes (1937)|8K', 1937, 'documentaries', 'newsreels', 300, 'hindenburg,zeppelin', False),
    ('revival-org', 'D1WD7sS637k', 'Revival of Organisms (1940)|8K', 1940, 'documentaries', 'science', 1200, 'soviet,science', False),
    ('wonderful-w', 'UtVxs89CUPc', 'Wonderful World (1959)|8K', 1959, 'documentaries', 'travelogues', 900, 'coca-cola,travel', False),
    ('golden-gate', '2Ua5qHZ5QDw', 'Golden Gate San Francisco|8K', 1950, 'documentaries', 'travelogues', 600, 'san francisco', False),
    ('popeye-mara', '3gzbxznJ_PM', 'Popeye Marathon 4h+|8K', 1945, 'cartoons', 'popeye', 16864, 'popeye,marathon', False),
    ('santas-surp', 'f8ZGvwk0k-o', 'Santas Surprise (1947)|8K', 1947, 'christmas', 'cartoons', 505, 'weihnachten', False),
    ('commodore64', 'j4r3bPPQza0', 'Commodore 64 Doku', 1982, 'documentaries', 'tech', 1712, 'commodore,c64', False),
    ('superman-11', '6A2_RKWP2X4', 'Superman 11th Hour (1942)', 1942, 'cartoons', 'superman', 540, 'superman,wwii', False),
    ('children-mv', 'EhZdQ74_sCM', 'Kinder machen Filme', 1960, 'documentaries', 'educational', 547, 'prelinger', False),
    ('grim-game', 'EMnokZOLpzU', 'Grim Game (1919) Houdini', 1919, 'classic-films', 'thriller', 244, 'houdini', False),
    ('candid-cand', 'ExbniPRgH70', 'Candid Candidate (1937)', 1937, 'cartoons', 'political', 364, 'politik,satire', False),
    ('charade', 'PSdJJaxI4gM', 'Charade (1953) Mystery', 1953, 'classic-films', 'mystery', 4901, 'mystery', False),
    ('caligari', '5WVJHELSD7A', 'Cabinet Dr. Caligari (1920)', 1920, 'classic-films', 'horror', 4518, 'expressionismus', False),
    ('white-zombie', 'd8Ak1R_eOlY', 'White Zombie (1932)', 1932, 'classic-films', 'horror', 3920, 'bela lugosi', False),
    ('felix-good', 'DMDp52B3Nv4', 'Felix macht es gut (1922)', 1922, 'cartoons', 'felix', 587, 'felix the cat', False),
    ('felix-love', 'eySdw7WrSps', 'Felix verliebt sich (1922)', 1922, 'cartoons', 'felix', 367, 'felix the cat', False),
    ('felix-fair', 'E7X_PG6wJEk', 'Felix auf Jahrmarkt (1922)', 1922, 'cartoons', 'felix', 264, 'felix the cat', False),
    ('big-game', '9_rhugQFh8w', 'Grosswildjaeger Cartoon', 1930, 'cartoons', 'adventure', 312, 'safari,jagd', False),
    ('felix-circus', 'kLpuqswC0IE', 'Felix im Zirkus (1921)', 1921, 'cartoons', 'felix', 300, 'felix the cat', False),
]

for i, v in enumerate(vids, 2):
    ws2.cell(row=i, column=1, value=i - 1).font = nf
    ws2.cell(row=i, column=1).border = tb
    for j, val in enumerate(v, 2):
        c = ws2.cell(row=i, column=j, value=val)
        c.font = nf
        c.border = tb
    for col in range(11, 16):
        c = ws2.cell(row=i, column=col, value='PRUEFEN')
        c.font = of
        c.border = tb

lr = len(vids) + 1
sr = lr + 2
ws2.cell(row=sr, column=1, value='STATISTIKEN').font = gf
ws2.cell(row=sr, column=1).fill = gold_fill

stats = [
    ('Gesamtanzahl', f'=COUNTA(B2:B{lr})'),
    ('Gesamtdauer (Stunden)', f'=SUM(H2:H{lr})/3600'),
    ('Kat: christmas', f'=COUNTIF(F2:F{lr},"christmas")'),
    ('Kat: classic-films', f'=COUNTIF(F2:F{lr},"classic-films")'),
    ('Kat: cartoons', f'=COUNTIF(F2:F{lr},"cartoons")'),
    ('Kat: comedy', f'=COUNTIF(F2:F{lr},"comedy")'),
    ('Kat: documentaries', f'=COUNTIF(F2:F{lr},"documentaries")'),
    ('Kat: propaganda', f'=COUNTIF(F2:F{lr},"propaganda")'),
    ('Kat: commercials', f'=COUNTIF(F2:F{lr},"commercials")'),
    ('Aeltestes (Jahr)', f'=MIN(E2:E{lr})'),
    ('Neuestes (Jahr)', f'=MAX(E2:E{lr})'),
]
for j, (label, formula) in enumerate(stats, sr + 1):
    ws2.cell(row=j, column=1, value=label).font = nf
    c = ws2.cell(row=j, column=2, value=formula)
    c.font = ff
    if 'Stunden' in label:
        c.number_format = '0.0'

# === SHEET 3-7: Quick sheets ===
sheets_data = {
    'YT_Optimierung': {
        'color': '00C853',
        'headers': ['Nr', 'Phase', 'Aufgabe', 'Prio', 'Status', 'Impact'],
        'data': [
            (1, 'P1', 'Playlist: Superman Fleischer', 'P0', 'TODO', 'Hoch'),
            (2, 'P1', 'Playlist: Popeye Classic', 'P0', 'TODO', 'Hoch'),
            (3, 'P1', 'Playlist: Chaplin Silent Comedy', 'P0', 'TODO', 'Hoch'),
            (4, 'P1', 'Playlist: Keaton Silent Comedy', 'P0', 'TODO', 'Hoch'),
            (5, 'P1', 'Playlist: Felix the Cat', 'P1', 'TODO', 'Mittel'),
            (6, 'P1', 'Playlist: Weihnachts-Specials', 'P0', 'TODO', 'Hoch'),
            (7, 'P1', 'Playlist: Horror-Klassiker', 'P0', 'TODO', 'Hoch'),
            (8, 'P1', 'Playlist: Stummfilm-Meisterwerke', 'P1', 'TODO', 'Mittel'),
            (9, 'P1', 'Playlist: Fritz Lang', 'P1', 'TODO', 'Mittel'),
            (10, 'P1', 'Playlist: WWII Dokus', 'P1', 'TODO', 'Mittel'),
            (11, 'P1', 'Playlist: Cold War', 'P1', 'TODO', 'Mittel'),
            (12, 'P1', 'Playlist: Early Animation', 'P2', 'TODO', 'Niedrig'),
            (13, 'P1', 'Playlist: 8K Restorations', 'P1', 'TODO', 'Hoch'),
            (14, 'P2', 'Chapters: Chaplin Film Fest', 'P0', 'TEILWEISE', 'Hoch'),
            (15, 'P2', 'Chapters: Popeye Marathon', 'P0', 'VORHANDEN', 'Hoch'),
            (16, 'P2', 'Chapters: Kirby Abridged', 'P1', 'TEILWEISE', 'Mittel'),
            (17, 'P3', 'Titel standardisieren alle Videos', 'P0', 'TODO', 'Hoch'),
            (18, 'P3', 'Beschreibungs-Template erstellen', 'P0', 'TODO', 'Hoch'),
            (19, 'P4', 'Channel-Homepage 12 Sections', 'P0', 'TODO', 'Hoch'),
            (20, 'P5', 'Thumbnail-Templates pro Kategorie', 'P1', 'TODO', 'Hoch'),
            (21, 'X', 'Multi-Language Metadaten EN/FR', 'P1', 'TODO', 'Hoch'),
            (22, 'X', 'Shorts-Strategie', 'P2', 'TODO', 'Hoch'),
            (23, 'X', 'End Screens einrichten', 'P1', 'TODO', 'Mittel'),
            (24, 'X', 'robots.txt AI Crawler', 'P0', 'TODO', 'Hoch'),
        ]
    },
    'Algo_Checkliste': {
        'color': 'FF9100',
        'headers': ['Signal', 'Aktuell', 'Ziel', 'Massnahme', 'Prio'],
        'data': [
            ('CTR', 'Unbekannt', '5-6%+', 'Thumbnails+Titel A/B Testing', 'P0'),
            ('AVD', 'Unbekannt', '50-60%', 'Chapters, bessere Intros', 'P0'),
            ('Satisfaction', 'Unbekannt', 'Positiv', 'Engagement, Kommentare', 'P1'),
            ('Playlist Sessions', '0 Playlists', '+22% Views', 'Series Playlists erstellen', 'P0'),
            ('Chapters', '1 von 3', 'Alle Kompilationen', 'Fehlende ergaenzen', 'P0'),
            ('Multi-Language', 'Nur DE', 'DE+EN+FR', 'Metadaten uebersetzen', 'P1'),
            ('Shorts', 'Keine', '2-3/Woche', 'Clips aus Long-Form', 'P2'),
            ('Upload-Rhythmus', 'Unregelmaessig', '1-2/Woche', 'Upload-Kalender', 'P1'),
            ('End Screens', 'Keine', 'Alle Videos', 'Einrichten', 'P1'),
            ('AI Crawler', 'Blockiert', 'Erlaubt', 'robots.txt anpassen', 'P0'),
            ('Website Schema', 'Teilweise', 'VideoObject pro Seite', 'JSON-LD ergaenzen', 'P0'),
        ]
    },
    'SEO_Audit': {
        'color': '2979FF',
        'headers': ['Seite', 'URL', 'Title', 'OG', 'Schema', 'hreflang', 'AI_OK', 'Score'],
        'data': [
            ('Homepage', '/', 'Ja', 'PRUEFEN', 'Org+FAQ+ItemList', 'Nein', 'Nein', 3),
            ('Timeline', '/timeline', 'PRUEFEN', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 2),
            ('Live', '/live', 'PRUEFEN', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 2),
            ('Video Detail', '/watch/:id', 'PRUEFEN', 'PRUEFEN', 'VideoObj?', 'Nein', 'Nein', 2),
            ('Wochenschau', '/wochenschau', '-', '-', '-', '-', '-', 0),
            ('Browse', '/browse', 'PRUEFEN', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 1),
            ('Search', '/search', 'PRUEFEN', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 1),
            ('Series', '/series', 'PRUEFEN', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 1),
            ('Impressum', '/impressum', 'Ja', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 2),
            ('Datenschutz', '/datenschutz', 'Ja', 'PRUEFEN', 'PRUEFEN', 'Nein', 'Nein', 2),
        ]
    },
    'Tech_Schulden': {
        'color': 'FF5252',
        'headers': ['ID', 'Beschreibung', 'Prio', 'Status', 'Impact'],
        'data': [
            ('TD-01', 'robots.txt: AI Crawler nicht erlaubt', 'P0', 'TODO', 'Hoch'),
            ('TD-02', 'Kein VideoObject Schema pro Video', 'P0', 'TODO', 'Hoch'),
            ('TD-03', 'Keine OG/Twitter Meta Tags pro Route', 'P0', 'TODO', 'Hoch'),
            ('TD-04', 'Kein hreflang DE/EN/FR', 'P1', 'TODO', 'Mittel'),
            ('TD-05', 'YouTube iframes ohne Facade (LCP)', 'P0', 'TODO', 'Hoch'),
            ('TD-06', 'Kein CDN konfiguriert', 'P1', 'TODO', 'Mittel'),
            ('TD-07', 'Test Coverage < 50%', 'P1', 'TODO', 'Mittel'),
            ('TD-08', 'i18n FR fast leer', 'P1', 'TODO', 'Mittel'),
            ('TD-09', 'Keine Video Sitemap mit video: Tags', 'P0', 'TODO', 'Hoch'),
            ('TD-10', 'WochenschauPage fehlt', 'P0', 'TODO', 'Hoch'),
            ('TD-11', 'SSR/Pre-rendering fehlt', 'P2', 'TODO', 'Hoch'),
            ('TD-12', 'Matomo Analytics nicht live', 'P1', 'TODO', 'Mittel'),
        ]
    },
    'i18n_Status': {
        'color': '7C4DFF',
        'headers': ['Bereich', 'DE', 'EN', 'FR'],
        'data': [
            ('Navigation', 'Voll', 'Voll', 'Teilw'),
            ('Footer', 'Voll', 'Voll', 'Teilw'),
            ('HomePage', 'Voll', 'Teilw', 'Fehl'),
            ('SearchPage', 'Teilw', 'Teilw', 'Fehl'),
            ('VideoPlayer', 'Teilw', 'Teilw', 'Fehl'),
            ('BrowsePage', 'Teilw', 'Teilw', 'Fehl'),
            ('TimelinePage', 'Teilw', 'Fehl', 'Fehl'),
            ('SeriesPage', 'Teilw', 'Fehl', 'Fehl'),
            ('Legal Pages', 'Voll', 'Fehl', 'Fehl'),
            ('CookieConsent', 'Voll', 'Voll', 'Fehl'),
            ('Kategorien', 'Voll', 'Teilw', 'Fehl'),
            ('WochenschauPage', 'Geplant', 'Geplant', 'Geplant'),
        ]
    }
}

for name, cfg in sheets_data.items():
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = cfg['color']
    hdrs = cfg['headers']
    for c, h in enumerate(hdrs, 1):
        ws.cell(row=1, column=c, value=h)
    sh(ws, 1, len(hdrs))
    for c in range(1, len(hdrs) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 15
    if len(hdrs) > 2:
        ws.column_dimensions['C' if len(hdrs) > 4 else 'B'].width = 45

    for i, row in enumerate(cfg['data'], 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = nf
            c.border = tb
            if val == 'TODO':
                c.fill = red_f
            elif str(val).startswith('TEIL') or val == 'PRUEFEN' or val == 'Teilw':
                c.fill = yellow_f
            elif str(val).startswith('VORH') or val == 'Ja' or val == 'Voll':
                c.fill = green_f
            elif val == 'Fehl':
                c.fill = red_f
            elif val == 'P0':
                c.font = rf
            elif val == 'Hoch':
                c.font = rf
            elif val == 'Nein':
                c.fill = red_f

    lr_s = len(cfg['data']) + 1
    ws.cell(row=lr_s + 2, column=1, value='Fortschritt').font = gf
    status_col = 'E' if len(hdrs) >= 5 else 'D'
    ws.cell(row=lr_s + 2, column=2, value=f'=COUNTIF({status_col}2:{status_col}{lr_s},"DONE")/COUNTA({status_col}2:{status_col}{lr_s})')
    ws.cell(row=lr_s + 2, column=2).font = ff
    ws.cell(row=lr_s + 2, column=2).number_format = '0.0%'

out = 'D:/remaike.TV/.claude/worktrees/exciting-golick/frai_tv_masterplan.xlsx'
wb.save(out)
print(f'XLSX erstellt: {out}')
print(f'Sheets: {wb.sheetnames}')
print(f'Videos im Katalog: {len(vids)}')
